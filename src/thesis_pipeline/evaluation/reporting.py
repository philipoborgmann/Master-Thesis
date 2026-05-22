"""Excel writer, CSV writer and console summary for the evaluation stage."""

from __future__ import annotations

from pathlib import Path
from typing import Mapping

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Style constants for the Excel sheets.
# ---------------------------------------------------------------------------

HEADER_FILL  = "1F4E78"   # dark blue
HEADER_FONT  = "FFFFFF"   # white
CATEGORY_COLORS: dict[str, str] = {
    "benchmark": "BFBFBF",   # grey
    "economic":  "BDD7EE",   # light blue
    "sentiment": "C6E0B4",   # light green
    "combined":  "FFD966",   # amber
}


# ---------------------------------------------------------------------------
# Leaderboard + thesis summary
# ---------------------------------------------------------------------------

def build_leaderboard(pooled: pd.DataFrame, top_n: int = 20) -> pd.DataFrame:
    if pooled.empty:
        return pooled
    cols = [c for c in ("horizon", "set_id", "category", "sentiment_model", "label",
                        "accuracy", "balanced_accuracy", "f1", "brier_score",
                        "log_loss", "n_obs", "n_tickers",
                        "mcnemar_pval_vs_b1", "significant_vs_b1",
                        "mcnemar_pval_vs_b2", "significant_vs_b2",
                        "mcnemar_pval", "significant_vs_benchmark") if c in pooled.columns]
    out = (pooled[cols]
           .sort_values(["horizon", "accuracy"], ascending=[True, False])
           .groupby("horizon", group_keys=False)
           .head(top_n)
           .reset_index(drop=True))
    return out


def build_summary(pooled: pd.DataFrame,
                  threshold_df: pd.DataFrame,
                  volatility_df: pd.DataFrame,
                  *,
                  threshold_lift: pd.DataFrame | None = None,
                  market_cap_df: pd.DataFrame | None = None,
                  interaction_df: pd.DataFrame | None = None,
                  economic_df: pd.DataFrame | None = None,
                  economic_threshold_df: pd.DataFrame | None = None,
                  extra_rows: list[dict] | None = None) -> pd.DataFrame:
    """Thesis-ready aggregate overview as a long-form table."""
    rows: list[dict] = []
    if not pooled.empty:
        # Average accuracy per category × horizon
        cat = (pooled.dropna(subset=["accuracy"])
                     .groupby(["horizon", "category"], dropna=False)["accuracy"]
                     .mean().round(6).reset_index())
        for _, r in cat.iterrows():
            rows.append({
                "section": "avg_accuracy_per_category",
                "horizon": r["horizon"], "category": r["category"],
                "metric": "accuracy", "value": r["accuracy"],
            })
        # Best sentiment model per horizon
        sent = pooled[pooled["category"].str.lower().isin(["sentiment", "combined"])]
        if not sent.empty:
            best = (sent.dropna(subset=["accuracy"])
                        .sort_values("accuracy", ascending=False)
                        .groupby("horizon").head(1))
            for _, r in best.iterrows():
                rows.append({
                    "section": "best_sentiment_model",
                    "horizon": r["horizon"],
                    "set_id":  r["set_id"],
                    "sentiment_model": r["sentiment_model"],
                    "metric": "accuracy", "value": r["accuracy"],
                })
        # Count of significant benchmark outperformance
        if "significant_vs_benchmark" in pooled.columns:
            sig = (pooled.groupby("horizon")["significant_vs_benchmark"]
                          .apply(lambda s: int(bool(s.fillna(False).sum())))
                          .reset_index())
            sig.columns = ["horizon", "n_significant"]
            for _, r in sig.iterrows():
                rows.append({
                    "section": "n_significant_vs_benchmark",
                    "horizon": r["horizon"],
                    "metric": "count", "value": int(r["n_significant"]),
                })
    if not volatility_df.empty:
        vol = (volatility_df.dropna(subset=["accuracy"])
                            .groupby(["horizon", "vol_regime"])["accuracy"]
                            .mean().round(6).reset_index())
        for _, r in vol.iterrows():
            rows.append({
                "section": "avg_accuracy_per_vol_regime",
                "horizon": r["horizon"], "vol_regime": r["vol_regime"],
                "metric": "accuracy", "value": r["accuracy"],
            })
    if not threshold_df.empty:
        thr = (threshold_df.dropna(subset=["accuracy"])
                           .groupby(["horizon", "threshold"])
                           .agg(accuracy=("accuracy", "mean"),
                                coverage=("coverage", "mean"))
                           .round(6).reset_index())
        for _, r in thr.iterrows():
            rows.append({
                "section": "threshold_tradeoff",
                "horizon": r["horizon"], "threshold": r["threshold"],
                "metric": "accuracy", "value": r["accuracy"],
                "coverage": r["coverage"],
            })
    if market_cap_df is not None and not market_cap_df.empty:
        small = (market_cap_df[market_cap_df["mcap_regime"] == "small"]
                 .dropna(subset=["accuracy"])
                 .sort_values("accuracy", ascending=False)
                 .groupby("horizon", group_keys=False).head(1))
        for _, r in small.iterrows():
            rows.append({
                "section": "best_small_cap_model",
                "horizon": r["horizon"], "set_id": r["set_id"],
                "sentiment_model": r.get("sentiment_model", "-"),
                "metric": "accuracy", "value": float(r["accuracy"]),
            })

    if interaction_df is not None and not interaction_df.empty:
        sub = (interaction_df[(interaction_df["mcap_regime"] == "small") &
                              (interaction_df["vol_regime"]  == "high")]
               .dropna(subset=["accuracy"])
               .sort_values("accuracy", ascending=False)
               .groupby("horizon", group_keys=False).head(1))
        for _, r in sub.iterrows():
            rows.append({
                "section": "best_small_high_vol_model",
                "horizon": r["horizon"], "set_id": r["set_id"],
                "sentiment_model": r.get("sentiment_model", "-"),
                "metric": "accuracy", "value": float(r["accuracy"]),
            })

    if economic_df is not None and not economic_df.empty:
        for metric, label in (("sharpe", "best_sharpe"),
                              ("cumulative_return", "best_cumulative_return")):
            valid = economic_df.dropna(subset=[metric])
            if valid.empty:
                continue
            gross = (valid[valid["cost_bps"] == 0]
                     .sort_values(metric, ascending=False)
                     .groupby("horizon", group_keys=False).head(1))
            for _, r in gross.iterrows():
                rows.append({
                    "section": label,
                    "horizon": r["horizon"], "set_id": r["set_id"],
                    "sentiment_model": r.get("sentiment_model", "-"),
                    "cost_bps": int(r["cost_bps"]),
                    "metric": metric, "value": float(r[metric]),
                })
            if metric == "sharpe":
                net = (valid[valid["cost_bps"] > 0]
                       .sort_values(metric, ascending=False)
                       .groupby("horizon", group_keys=False).head(1))
                for _, r in net.iterrows():
                    rows.append({
                        "section": "best_net_sharpe",
                        "horizon": r["horizon"], "set_id": r["set_id"],
                        "sentiment_model": r.get("sentiment_model", "-"),
                        "cost_bps": int(r["cost_bps"]),
                        "metric": metric, "value": float(r[metric]),
                    })

    if economic_threshold_df is not None and not economic_threshold_df.empty:
        valid = economic_threshold_df.dropna(subset=["sharpe"])
        if not valid.empty:
            best = (valid.sort_values("sharpe", ascending=False)
                         .groupby(["horizon", "cost_bps"], group_keys=False).head(1))
            for _, r in best.iterrows():
                rows.append({
                    "section":  "best_threshold_sharpe",
                    "horizon":  r["horizon"], "set_id": r["set_id"],
                    "sentiment_model": r.get("sentiment_model", "-"),
                    "threshold": float(r["threshold"]),
                    "cost_bps":  int(r["cost_bps"]),
                    "metric":   "sharpe", "value": float(r["sharpe"]),
                })

    if extra_rows:
        rows.extend(extra_rows)

    if threshold_lift is not None and not threshold_lift.empty:
        # Best matched-observation lift per (horizon, threshold, benchmark).
        lift = (threshold_lift.dropna(subset=["lift_accuracy"])
                              .sort_values("lift_accuracy", ascending=False)
                              .groupby(["horizon", "threshold", "benchmark"])
                              .head(1))
        for _, r in lift.iterrows():
            rows.append({
                "section": "best_threshold_lift",
                "horizon": r["horizon"], "threshold": r["threshold"],
                "benchmark": r["benchmark"],
                "set_id":    r["set_id"],
                "sentiment_model": r.get("sentiment_model", "-"),
                "metric": "lift_accuracy", "value": r["lift_accuracy"],
                "n_matched": int(r.get("n_matched", 0)),
                "significant": bool(r.get("significant_vs_benchmark", False)),
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Excel formatting helpers
# ---------------------------------------------------------------------------

def _style_header(ws) -> None:
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    font = Font(color=HEADER_FONT, bold=True)
    align = Alignment(horizontal="left", vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _freeze_and_filter(ws) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def _category_fills(ws, category_col_letter: str | None) -> None:
    if category_col_letter is None:
        return
    from openpyxl.styles import PatternFill
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{category_col_letter}{row}"]
        val = str(cell.value).lower() if cell.value is not None else ""
        color = CATEGORY_COLORS.get(val)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def _conditional_format_accuracy(ws, header_to_col: Mapping[str, str]) -> None:
    col_letter = header_to_col.get("accuracy")
    if col_letter is None or ws.max_row < 2:
        return
    from openpyxl.formatting.rule import ColorScaleRule
    rule = ColorScaleRule(
        start_type="num", start_value=0.40, start_color="F8696B",
        mid_type="num",   mid_value=0.50,   mid_color="FFEB84",
        end_type="num",   end_value=0.60,   end_color="63BE7B",
    )
    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
    ws.conditional_formatting.add(rng, rule)


def _write_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    if df is None:
        df = pd.DataFrame()
    if df.empty:
        df = pd.DataFrame({"info": [f"no data for sheet '{sheet_name}'"]})
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.book[sheet_name]
    _style_header(ws)
    _freeze_and_filter(ws)
    header_to_col = {cell.value: cell.column_letter for cell in ws[1]}
    _category_fills(ws, header_to_col.get("category"))
    _conditional_format_accuracy(ws, header_to_col)


# ---------------------------------------------------------------------------
# Public writers
# ---------------------------------------------------------------------------

def write_csv_outputs(out_dir: Path, *,
                      pooled: pd.DataFrame,
                      per_ticker: pd.DataFrame,
                      threshold: pd.DataFrame,
                      volatility: pd.DataFrame,
                      threshold_lift: pd.DataFrame | None = None,
                      mcnemar: pd.DataFrame | None = None,
                      market_cap: pd.DataFrame | None = None,
                      regime_interaction: pd.DataFrame | None = None,
                      economic: pd.DataFrame | None = None,
                      economic_threshold: pd.DataFrame | None = None,
                      economic_by_ticker: pd.DataFrame | None = None
                      ) -> dict[str, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "pooled":    out_dir / "pooled_metrics.csv",
        "per_ticker": out_dir / "per_ticker_metrics.csv",
        "threshold":  out_dir / "threshold_analysis.csv",
        "volatility": out_dir / "volatility_stratification.csv",
    }
    pooled.to_csv(paths["pooled"], index=False)
    per_ticker.to_csv(paths["per_ticker"], index=False)
    threshold.to_csv(paths["threshold"], index=False)
    volatility.to_csv(paths["volatility"], index=False)
    if threshold_lift is not None:
        paths["threshold_lift"] = out_dir / "threshold_lift.csv"
        threshold_lift.to_csv(paths["threshold_lift"], index=False)
    if mcnemar is not None:
        paths["mcnemar"] = out_dir / "mcnemar_tests.csv"
        mcnemar.to_csv(paths["mcnemar"], index=False)
    if market_cap is not None:
        paths["market_cap"] = out_dir / "market_cap_stratification.csv"
        market_cap.to_csv(paths["market_cap"], index=False)
    if regime_interaction is not None:
        paths["regime_interaction"] = out_dir / "regime_interaction.csv"
        regime_interaction.to_csv(paths["regime_interaction"], index=False)
    if economic is not None:
        paths["economic"] = out_dir / "economic_performance.csv"
        economic.to_csv(paths["economic"], index=False)
    if economic_threshold is not None:
        paths["economic_threshold"] = out_dir / "economic_performance_by_threshold.csv"
        economic_threshold.to_csv(paths["economic_threshold"], index=False)
    if economic_by_ticker is not None:
        paths["economic_by_ticker"] = out_dir / "economic_performance_by_ticker.csv"
        economic_by_ticker.to_csv(paths["economic_by_ticker"], index=False)
    return paths


def write_excel_report(out_path: Path, *,
                       pooled: pd.DataFrame,
                       per_ticker: pd.DataFrame,
                       threshold: pd.DataFrame,
                       volatility: pd.DataFrame,
                       leaderboard: pd.DataFrame,
                       summary: pd.DataFrame,
                       threshold_lift: pd.DataFrame | None = None,
                       mcnemar: pd.DataFrame | None = None,
                       market_cap: pd.DataFrame | None = None,
                       regime_interaction: pd.DataFrame | None = None,
                       economic: pd.DataFrame | None = None,
                       economic_threshold: pd.DataFrame | None = None,
                       economic_by_ticker: pd.DataFrame | None = None) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        _write_sheet(writer, "pooled_metrics",                pooled)
        _write_sheet(writer, "per_ticker_metrics",            per_ticker)
        _write_sheet(writer, "threshold_analysis",            threshold)
        if threshold_lift is not None:
            _write_sheet(writer, "threshold_lift",            threshold_lift)
        _write_sheet(writer, "volatility_stratification",     volatility)
        if market_cap is not None:
            _write_sheet(writer, "market_cap_stratification", market_cap)
        if regime_interaction is not None:
            _write_sheet(writer, "regime_interaction",        regime_interaction)
        if economic is not None:
            _write_sheet(writer, "economic_performance",      economic)
        if economic_threshold is not None:
            _write_sheet(writer, "economic_by_threshold",     economic_threshold)
        if economic_by_ticker is not None:
            _write_sheet(writer, "economic_by_ticker",        economic_by_ticker)
        if mcnemar is not None:
            _write_sheet(writer, "mcnemar_tests",             mcnemar)
        _write_sheet(writer, "leaderboard",                   leaderboard)
        _write_sheet(writer, "summary",                       summary)
    return out_path


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------

def print_console_summary(*,
                          pooled: pd.DataFrame,
                          threshold: pd.DataFrame,
                          mcnemar: pd.DataFrame,
                          threshold_lift: pd.DataFrame | None = None,
                          market_cap: pd.DataFrame | None = None,
                          regime_interaction: pd.DataFrame | None = None,
                          economic: pd.DataFrame | None = None,
                          economic_threshold: pd.DataFrame | None = None) -> None:
    print("\n" + "=" * 72)
    print(" SIGNAL EVALUATION SUMMARY")
    print("=" * 72)
    if pooled.empty:
        print("  (no pooled metrics available)")
        return

    # Benchmark accuracy per horizon
    bench = pooled[pooled["set_id"] == "B1"][["horizon", "accuracy"]].sort_values("horizon")
    if not bench.empty:
        print("\n  Benchmark (B1) accuracy by horizon:")
        for _, r in bench.iterrows():
            acc = r["accuracy"]
            acc_str = f"{acc:.4f}" if pd.notna(acc) else "n/a"
            print(f"    {r['horizon']:>3s}  acc={acc_str}")

    # Top-5 per horizon
    print("\n  Top-5 by accuracy per horizon:")
    for hz, grp in pooled.dropna(subset=["accuracy"]).groupby("horizon"):
        top = grp.sort_values("accuracy", ascending=False).head(5)
        print(f"   [{hz}]")
        for _, r in top.iterrows():
            sm = r.get("sentiment_model", "-")
            tag = f"/{sm}" if sm and sm != "-" else ""
            print(f"     {r['set_id']}{tag:14s} acc={r['accuracy']:.4f}  "
                  f"f1={r.get('f1', float('nan')):.4f}  "
                  f"brier={r.get('brier_score', float('nan')):.4f}  "
                  f"n={int(r.get('n_obs', 0))}")

    # Significant outperformers — broken out per benchmark.
    if not mcnemar.empty and "significant_vs_benchmark" in mcnemar.columns:
        for bid, grp in mcnemar.groupby("benchmark"):
            n_sig = int(grp["significant_vs_benchmark"].fillna(False).sum())
            print(f"\n  Significant (α=0.05) outperformers vs {bid}: {n_sig}")
            sig = grp[grp["significant_vs_benchmark"]].sort_values("mcnemar_pval")
            for _, r in sig.head(10).iterrows():
                sm = r.get("sentiment_model", "-")
                tag = f"/{sm}" if sm and sm != "-" else ""
                print(f"    {r['horizon']:>3s}  {r['set_id']}{tag:14s} "
                      f"stat={r['mcnemar_stat']:.3f}  p={r['mcnemar_pval']:.4f}")

    # Best sentiment model per horizon
    sent = pooled[pooled["category"].str.lower().isin(["sentiment", "combined"])]
    if not sent.empty:
        print("\n  Best sentiment/combined set per horizon:")
        best = (sent.dropna(subset=["accuracy"])
                    .sort_values("accuracy", ascending=False)
                    .groupby("horizon").head(1))
        for _, r in best.iterrows():
            print(f"    {r['horizon']:>3s}  {r['set_id']}/{r['sentiment_model']:8s} "
                  f"acc={r['accuracy']:.4f}")

    # Threshold highlights
    if not threshold.empty:
        print("\n  Threshold tradeoffs (mean across sets):")
        thr = (threshold.dropna(subset=["accuracy"])
                        .groupby(["horizon", "threshold"])
                        .agg(accuracy=("accuracy", "mean"),
                             coverage=("coverage", "mean"))
                        .reset_index())
        for hz, grp in thr.groupby("horizon"):
            print(f"   [{hz}]")
            for _, r in grp.iterrows():
                print(f"     t={r['threshold']:.2f}  acc={r['accuracy']:.4f}  "
                      f"coverage={r['coverage']:.3f}")

    # Threshold lift — does sentiment beat the benchmark on the model's
    # high-conviction observations?
    if threshold_lift is not None and not threshold_lift.empty:
        print("\n  Top sentiment lifts on matched obs (model traded → vs benchmark):")
        for hz, hz_grp in threshold_lift.dropna(subset=["lift_accuracy"]) \
                                        .groupby("horizon"):
            print(f"   [{hz}]")
            top = (hz_grp.sort_values("lift_accuracy", ascending=False)
                          .head(5))
            for _, r in top.iterrows():
                sm = r.get("sentiment_model", "-")
                tag = f"/{sm}" if sm and sm != "-" else ""
                star = "*" if bool(r.get("significant_vs_benchmark", False)) else " "
                print(f"     {star} {r['set_id']}{tag:14s} t={r['threshold']:.2f} "
                      f"vs {r['benchmark']}  lift={r['lift_accuracy']:+.4f} "
                      f"(model={r['accuracy_model']:.4f}, "
                      f"bench={r['accuracy_benchmark']:.4f}, "
                      f"n={int(r['n_matched'])})")

    # Market-cap stratification — strongest small-cap lift.
    if market_cap is not None and not market_cap.empty:
        print("\n  Strongest small-cap accuracy per horizon:")
        small = (market_cap[market_cap["mcap_regime"] == "small"]
                 .dropna(subset=["accuracy"])
                 .sort_values("accuracy", ascending=False)
                 .groupby("horizon", group_keys=False).head(3))
        for hz, grp in small.groupby("horizon"):
            print(f"   [{hz}]")
            for _, r in grp.iterrows():
                sm = r.get("sentiment_model", "-")
                tag = f"/{sm}" if sm and sm != "-" else ""
                print(f"     {r['set_id']}{tag:14s} acc={r['accuracy']:.4f}  "
                      f"brier={r.get('brier_score', float('nan')):.4f}  "
                      f"n={int(r.get('n_obs', 0))}")

    # Small-cap × high-vol interaction — central thesis hypothesis.
    if regime_interaction is not None and not regime_interaction.empty:
        sub = (regime_interaction[(regime_interaction["mcap_regime"] == "small") &
                                  (regime_interaction["vol_regime"]  == "high")]
               .dropna(subset=["accuracy"])
               .sort_values("accuracy", ascending=False))
        if not sub.empty:
            print("\n  Best small-cap × high-vol accuracy per horizon:")
            top = sub.groupby("horizon", group_keys=False).head(3)
            for hz, grp in top.groupby("horizon"):
                print(f"   [{hz}]")
                for _, r in grp.iterrows():
                    sm = r.get("sentiment_model", "-")
                    tag = f"/{sm}" if sm and sm != "-" else ""
                    print(f"     {r['set_id']}{tag:14s} acc={r['accuracy']:.4f}  "
                          f"n={int(r.get('n_obs', 0))}  "
                          f"days={int(r.get('n_days', 0))}")

    # Economic — best Sharpe / cumulative return per horizon and cost.
    if economic is not None and not economic.empty:
        print("\n  Best Sharpe per (horizon × cost_bps):")
        valid = economic.dropna(subset=["sharpe"])
        if not valid.empty:
            best = (valid.sort_values("sharpe", ascending=False)
                         .groupby(["horizon", "cost_bps"], group_keys=False).head(1))
            for _, r in best.iterrows():
                sm = r.get("sentiment_model", "-")
                tag = f"/{sm}" if sm and sm != "-" else ""
                print(f"     {r['horizon']:>3s}  cost={int(r['cost_bps'])}bps  "
                      f"{r['set_id']}{tag:14s} "
                      f"sharpe={r['sharpe']:+.3f}  "
                      f"cum_ret={r.get('cumulative_return', float('nan')):+.3f}  "
                      f"max_dd={r.get('max_drawdown', float('nan')):.3f}  "
                      f"turnover={r.get('turnover', float('nan')):.3f}")

    if economic_threshold is not None and not economic_threshold.empty:
        print("\n  Best Sharpe per (horizon × threshold × cost_bps):")
        valid = economic_threshold.dropna(subset=["sharpe"])
        if not valid.empty:
            best = (valid.sort_values("sharpe", ascending=False)
                         .groupby(["horizon", "threshold", "cost_bps"],
                                  group_keys=False).head(1))
            for _, r in best.iterrows():
                sm = r.get("sentiment_model", "-")
                tag = f"/{sm}" if sm and sm != "-" else ""
                print(f"     {r['horizon']:>3s}  t={r['threshold']:.2f}  "
                      f"cost={int(r['cost_bps'])}bps  "
                      f"{r['set_id']}{tag:14s} "
                      f"sharpe={r['sharpe']:+.3f}  "
                      f"cov={r.get('coverage', float('nan')):.3f}  "
                      f"n_trades={int(r.get('n_trades', 0))}")
    print()
