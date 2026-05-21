#!/usr/bin/env python3
"""
Sentiment_Data_Load.py
======================
Loads raw Reddit sentiment data, cleans duplicates, maps subreddits to coin
tickers via an Excel lookup table, preprocesses text for sentiment scoring,
computes descriptive statistics, and exports processed data + statistics +
visualizations.

Input:
    - Data/Raw/Sentiment/{subreddit_name}/submission.csv
    - subreddit_ticker_mapping.xlsx

Output:
    - Data/Processed/Sentiment/sentiment_combined.csv
    - Outputs/deskriptiv/descriptive_statistics.xlsx
    - Outputs/deskriptiv/plots/*.png

Requirements:
    pip install pandas openpyxl matplotlib seaborn

Usage:
    python Sentiment_Data_Load.py
    python Sentiment_Data_Load.py --mapping ./subreddit_ticker_mapping.xlsx
"""

import argparse
import glob
import html
import os
import re
import sys

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════════════════
# 1. CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

KEEP_COLUMNS = [
    "submission",
    "subreddit",
    "created",
    "removed",
    "upvote_ratio",
    "score",
    "total_awards_received",
    "num_comments",
    "num_crossposts",
    "selftext",
    "title",
]

sns.set_theme(style="whitegrid", font_scale=1.1)
PALETTE = sns.color_palette("tab20", 30)

# ══════════════════════════════════════════════════════════════════════════════
# 2. LOAD TICKER MAPPING FROM EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def load_ticker_mapping(mapping_path: str) -> dict:
    df = pd.read_excel(mapping_path, sheet_name="Subreddit_Ticker_Mapping")
    df = df[df["included"].str.lower() == "yes"]
    mapping = dict(zip(df["subreddit"], df["ticker"]))
    print(f"[INFO] Loaded ticker mapping: {len(mapping)} subreddits -> "
          f"{len(set(mapping.values()))} unique tickers")
    return mapping

# ══════════════════════════════════════════════════════════════════════════════
# 3. LOAD AND COMBINE RAW DATA
# ══════════════════════════════════════════════════════════════════════════════

def load_raw_data(raw_dir: str, mapping: dict) -> pd.DataFrame:
    frames = []
    found_subs = []
    missing_subs = []

    for subreddit, ticker in sorted(mapping.items()):
        csv_path = os.path.join(raw_dir, subreddit, "submission.csv")
        if not os.path.exists(csv_path):
            missing_subs.append(subreddit)
            continue

        df = pd.read_csv(
            csv_path,
            usecols=lambda c: c in KEEP_COLUMNS,
            dtype={"submission": str, "subreddit": str,
                   "selftext": str, "title": str},
        )
        df["ticker"] = ticker
        frames.append(df)
        found_subs.append(subreddit)

    print(f"[INFO] Loaded {len(found_subs)} subreddit CSVs")
    if missing_subs:
        print(f"[WARN] Missing CSVs for: {', '.join(missing_subs)}")

    combined = pd.concat(frames, ignore_index=True)
    print(f"[INFO] Raw combined shape: {combined.shape[0]:,} rows x {combined.shape[1]} columns")
    return combined

# ══════════════════════════════════════════════════════════════════════════════
# 4. CLEAN DATA: DEDUPLICATE, RENAME, FORMAT
# ══════════════════════════════════════════════════════════════════════════════

def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    n_before = len(df)

    df = df.drop_duplicates(subset=["submission", "subreddit"], keep="first")
    n_deduped = n_before - len(df)
    print(f"[INFO] Duplicates removed: {n_deduped:,} "
          f"({n_deduped / n_before * 100:.2f}%)")

    df["created"] = pd.to_datetime(df["created"], unit="s", utc=True)
    df = df.rename(columns={"created": "date"})

    df["date_only"] = df["date"].dt.date
    df["hour"]      = df["date"].dt.hour
    df["weekday"]   = df["date"].dt.day_name()
    df["month"]     = df["date"].dt.to_period("M").astype(str)

    df = df.drop(columns=["submission"])
    df = df.sort_values("date").reset_index(drop=True)

    print(f"[INFO] Cleaned shape: {len(df):,} rows x {df.shape[1]} columns")
    print(f"[INFO] Date range: {df['date_only'].min()} -> {df['date_only'].max()}")
    return df

# ══════════════════════════════════════════════════════════════════════════════
# 5. TEXT PREPROCESSING
# ══════════════════════════════════════════════════════════════════════════════

# Compiled regex patterns (compiled once for performance on 479K+ rows)
_RE_URL = re.compile(r"https?://\S+|www\.\S+", re.IGNORECASE)
_RE_REDDIT_USER = re.compile(r"/?u/\w+")
_RE_REDDIT_SUB  = re.compile(r"/?r/\w+")
_RE_MARKDOWN_HEADERS = re.compile(r"^#{1,6}\s+", re.MULTILINE)
_RE_MARKDOWN_BOLD_ITALIC = re.compile(r"(\*{1,3}|_{1,3})(.*?)\1")
_RE_MARKDOWN_STRIKETHROUGH = re.compile(r"~~(.*?)~~")
_RE_MARKDOWN_LINKS = re.compile(r"\[([^\]]*)\]\([^\)]*\)")
_RE_MARKDOWN_BLOCKQUOTE = re.compile(r"^>\s?", re.MULTILINE)
_RE_MARKDOWN_CODE_BLOCK = re.compile(r"```.*?```", re.DOTALL)
_RE_MARKDOWN_INLINE_CODE = re.compile(r"`([^`]*)`")
_RE_WHITESPACE = re.compile(r"\s+")


def preprocess_text(text: str) -> str:
    """
    Lightweight preprocessing for Reddit post text, designed to work
    across VADER, FinBERT, and CryptoBERT without damaging any model's
    specific signal channels.

    Applied:
        1. Decode HTML entities (&amp; -> &, &lt; -> <)
        2. Remove URLs (http/https/www links)
        3. Remove Reddit-specific syntax (u/user, r/subreddit)
        4. Strip markdown formatting (keep underlying text)
        5. Collapse whitespace

    Preserved (intentionally NOT removed):
        - Emojis and emoticons (VADER scores them)
        - Capitalization (VADER uses ALL CAPS as intensity amplifier)
        - Punctuation including !, ?, ... (VADER uses as modifiers)
        - Stopwords (VADER uses negation: "not good"; transformers need context)
        - Crypto slang and ticker symbols ($BTC, HODL, etc.)
    """
    if not isinstance(text, str) or not text.strip():
        return text

    # 1. Decode HTML entities
    text = html.unescape(text)

    # 2. Remove URLs
    text = _RE_URL.sub("", text)

    # 3. Remove Reddit-specific syntax
    text = _RE_REDDIT_USER.sub("", text)
    text = _RE_REDDIT_SUB.sub("", text)

    # 4. Strip markdown (preserve underlying text content)
    text = _RE_MARKDOWN_CODE_BLOCK.sub("", text)
    text = _RE_MARKDOWN_INLINE_CODE.sub(r"\1", text)
    text = _RE_MARKDOWN_LINKS.sub(r"\1", text)
    text = _RE_MARKDOWN_HEADERS.sub("", text)
    text = _RE_MARKDOWN_BOLD_ITALIC.sub(r"\2", text)
    text = _RE_MARKDOWN_STRIKETHROUGH.sub(r"\1", text)
    text = _RE_MARKDOWN_BLOCKQUOTE.sub("", text)

    # 5. Collapse whitespace
    text = _RE_WHITESPACE.sub(" ", text).strip()

    return text


def apply_preprocessing(df: pd.DataFrame) -> pd.DataFrame:
    """
    Applies preprocess_text() to both title and selftext columns.
    Reports character reduction statistics.
    """
    print(f"\n[INFO] Preprocessing {len(df):,} posts...")

    # Title
    title_len_before = df["title"].str.len().mean()
    df["title"] = df["title"].apply(preprocess_text)
    title_len_after = df["title"].str.len().mean()
    title_pct = (1 - title_len_after / title_len_before) * 100

    # Selftext
    has_st = df["selftext"].notna() & (df["selftext"].str.len() > 0)
    st_len_before = df.loc[has_st, "selftext"].str.len().mean()
    df["selftext"] = df["selftext"].apply(preprocess_text)
    st_len_after = df.loc[has_st, "selftext"].str.len().mean()
    st_pct = (1 - st_len_after / st_len_before) * 100

    print(f"  Title:    avg length {title_len_before:.0f} -> {title_len_after:.0f} chars "
          f"({title_pct:.1f}% reduction)")
    print(f"  Selftext: avg length {st_len_before:.0f} -> {st_len_after:.0f} chars "
          f"({st_pct:.1f}% reduction)")

    return df

# ══════════════════════════════════════════════════════════════════════════════
# 6. DESCRIPTIVE STATISTICS
# ══════════════════════════════════════════════════════════════════════════════

def compute_descriptive_stats(df: pd.DataFrame) -> dict:
    stats = {}
    total_days = (df["date_only"].max() - df["date_only"].min()).days + 1
    total_hours = total_days * 24

    overall = pd.DataFrame([{
        "total_posts":        len(df),
        "unique_tickers":     df["ticker"].nunique(),
        "unique_subreddits":  df["subreddit"].nunique(),
        "date_start":         str(df["date_only"].min()),
        "date_end":           str(df["date_only"].max()),
        "total_days":         total_days,
        "avg_posts_per_day":  round(len(df) / total_days, 1),
        "avg_posts_per_hour": round(len(df) / total_hours, 2),
        "total_removed":      int((df["removed"] == 1).sum()),
        "removed_pct":        round((df["removed"] == 1).mean() * 100, 2),
    }]).T
    overall.columns = ["value"]
    overall.index.name = "metric"
    stats["overall_summary"] = overall

    coin_stats = (
        df.groupby("ticker")
          .agg(
              total_posts       = ("ticker", "size"),
              subreddits        = ("subreddit", "nunique"),
              subreddit_list    = ("subreddit", lambda x: ", ".join(sorted(x.unique()))),
              avg_posts_per_day = ("date_only", lambda x: round(len(x) / total_days, 1)),
              avg_posts_per_hour= ("date_only", lambda x: round(len(x) / total_hours, 2)),
              removed_count     = ("removed", lambda x: (x == 1).sum()),
              removed_pct       = ("removed", lambda x: round((x == 1).mean() * 100, 2)),
              avg_score         = ("score", "mean"),
              median_score      = ("score", "median"),
              avg_upvote_ratio  = ("upvote_ratio", "mean"),
              avg_comments      = ("num_comments", "mean"),
              avg_awards        = ("total_awards_received", "mean"),
          )
          .sort_values("total_posts", ascending=False)
          .round(2)
    )
    stats["per_coin"] = coin_stats

    monthly = df.groupby(["month", "ticker"]).size().unstack(fill_value=0)
    monthly["total"] = monthly.sum(axis=1)
    stats["monthly_volume"] = monthly

    hourly = df.groupby("hour").agg(
        total_posts = ("hour", "size"),
        avg_per_day = ("date_only", lambda x: round(len(x) / total_days, 2)),
    )
    stats["hourly_distribution"] = hourly

    day_order = ["Monday", "Tuesday", "Wednesday", "Thursday",
                 "Friday", "Saturday", "Sunday"]
    weekday = (
        df.groupby("weekday")
          .agg(
              total_posts  = ("weekday", "size"),
              avg_per_day  = ("date_only", lambda x: round(len(x) / (total_days / 7), 1)),
              avg_score    = ("score", "mean"),
              avg_comments = ("num_comments", "mean"),
          )
          .reindex(day_order)
          .round(2)
    )
    stats["weekday_distribution"] = weekday

    engagement = (
        df.groupby("ticker")
          .agg(
              avg_score          = ("score", "mean"),
              std_score          = ("score", "std"),
              max_score          = ("score", "max"),
              avg_upvote_ratio   = ("upvote_ratio", "mean"),
              avg_comments       = ("num_comments", "mean"),
              max_comments       = ("num_comments", "max"),
              avg_crossposts     = ("num_crossposts", "mean"),
              avg_awards         = ("total_awards_received", "mean"),
              pct_with_selftext  = ("selftext", lambda x: round(
                  x.notna().mean() * 100
                  - x.isin(["[removed]", "[deleted]", ""]).mean() * 100, 2
              )),
          )
          .sort_values("avg_score", ascending=False)
          .round(2)
    )
    stats["engagement"] = engagement

    print(f"\n{'─' * 60}")
    print("DESCRIPTIVE STATISTICS SUMMARY")
    print(f"{'─' * 60}")
    print(f"Total posts:         {len(df):>10,}")
    print(f"Unique tickers:      {df['ticker'].nunique():>10}")
    print(f"Date range:          {df['date_only'].min()} -> {df['date_only'].max()}")
    print(f"Avg posts/day:       {len(df) / total_days:>10.1f}")
    print(f"Avg posts/hour:      {len(df) / total_hours:>10.2f}")
    print(f"Removed posts:       {(df['removed'] == 1).sum():>10,} "
          f"({(df['removed'] == 1).mean() * 100:.1f}%)")
    print(f"\nTop 5 coins by volume:")
    for ticker, row in coin_stats.head().iterrows():
        print(f"  {ticker:6s} {int(row['total_posts']):>8,} posts  "
              f"({row['avg_posts_per_day']:.1f}/day, "
              f"{row['avg_posts_per_hour']:.2f}/hour)")
    print(f"{'─' * 60}")

    return stats

# ══════════════════════════════════════════════════════════════════════════════
# 7. EXPORT DESCRIPTIVE STATISTICS TO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def export_stats_to_excel(stats: dict, output_path: str):
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in stats.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        df_out = df.reset_index() if df.index.name or df.index.names[0] else df

        for col_idx, col_name in enumerate(df_out.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        for row_idx, row in enumerate(df_out.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx,
                               value=val if not isinstance(val, (np.integer, np.floating))
                               else float(val))
                cell.font = data_font
                cell.border = thin_border

        for col_idx, col_name in enumerate(df_out.columns, 1):
            max_len = max(
                len(str(col_name)),
                *(len(str(ws.cell(row=r, column=col_idx).value or ""))
                  for r in range(2, ws.max_row + 1))
            ) if ws.max_row > 1 else len(str(col_name))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                min(max_len + 3, 40)
            )
        ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"[INFO] Descriptive statistics -> {output_path}")

# ══════════════════════════════════════════════════════════════════════════════
# 8. VISUALIZATIONS
# ══════════════════════════════════════════════════════════════════════════════

def create_visualizations(df: pd.DataFrame, stats: dict, plot_dir: str):
    os.makedirs(plot_dir, exist_ok=True)
    coin_stats = stats["per_coin"]

    fig, ax = plt.subplots(figsize=(12, 6))
    data = coin_stats["total_posts"].sort_values(ascending=True)
    colors = plt.cm.Blues(np.linspace(0.3, 0.9, len(data)))
    data.plot.barh(ax=ax, color=colors)
    ax.set_xlabel("Total Posts")
    ax.set_title("Total Reddit Posts per Coin (2022)")
    ax.bar_label(ax.containers[0], fmt="{:,.0f}", padding=4, fontsize=8)
    plt.tight_layout()
    fig.savefig(os.path.join(plot_dir, "01_posts_per_coin.png"), dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(14, 5))
    daily = df.groupby("date_only").size()
    daily.index = pd.to_datetime(daily.index)
    daily_smooth = daily.rolling(7, center=True).mean()
    ax.fill_between(daily.index, daily.values, alpha=0.2, color="steelblue")
    ax.plot(daily_smooth.index, daily_smooth.values,
            color="steelblue", linewidth=1.5, label="7-day rolling avg")
    ax.set_ylabel("Posts per Day")
    ax.set_title("Daily Reddit Post Volume (All Coins)")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.legend()
    plt.tight_layout()
    fig.savefig(os.path.join(plot_dir, "02_daily_volume.png"), dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(14, 6))
    top_coins = coin_stats["total_posts"].nlargest(8).index.tolist()
    monthly_top = (
        df[df["ticker"].isin(top_coins)]
        .groupby(["month", "ticker"]).size().unstack(fill_value=0)
    )
    monthly_top.plot.bar(stacked=True, ax=ax, width=0.8, colormap="tab20")
    ax.set_ylabel("Posts")
    ax.set_title("Monthly Post Volume by Coin (Top 8)")
    ax.legend(title="Ticker", bbox_to_anchor=(1.02, 1), loc="upper left")
    ax.set_xlabel("")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    fig.savefig(os.path.join(plot_dir, "03_monthly_by_coin.png"), dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(10, 5))
    hourly = df.groupby("hour").size()
    colors_h = plt.cm.coolwarm(np.linspace(0, 1, 24))
    hourly.plot.bar(ax=ax, color=colors_h, edgecolor="white")
    ax.set_xlabel("Hour of Day (UTC)")
    ax.set_ylabel("Total Posts")
    ax.set_title("Posting Activity by Hour of Day (UTC)")
    plt.xticks(rotation=0)
    plt.tight_layout()
    fig.savefig(os.path.join(plot_dir, "04_hourly_distribution.png"), dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(10, 5))
    day_order = ["Monday", "Tuesday", "Wednesday", "Thursday",
                 "Friday", "Saturday", "Sunday"]
    weekday = df.groupby("weekday").size().reindex(day_order)
    weekday.plot.bar(ax=ax, color="steelblue", edgecolor="white")
    ax.set_xlabel("")
    ax.set_ylabel("Total Posts")
    ax.set_title("Posting Activity by Weekday")
    plt.xticks(rotation=30, ha="right")
    plt.tight_layout()
    fig.savefig(os.path.join(plot_dir, "05_weekday_distribution.png"), dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(12, 6))
    removal = coin_stats["removed_pct"].sort_values(ascending=True)
    colors_r = plt.cm.RdYlGn_r(removal.values / removal.max())
    removal.plot.barh(ax=ax, color=colors_r)
    ax.set_xlabel("Removed Posts (%)")
    ax.set_title("Moderation Removal Rate by Coin")
    ax.bar_label(ax.containers[0], fmt="{:.1f}%", padding=4, fontsize=8)
    plt.tight_layout()
    fig.savefig(os.path.join(plot_dir, "06_removal_rate.png"), dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(10, 7))
    eng = stats["engagement"].reset_index()
    scatter = ax.scatter(
        eng["avg_score"], eng["avg_comments"],
        s=coin_stats.loc[eng["ticker"], "total_posts"].values / 50,
        alpha=0.7, c=range(len(eng)), cmap="tab20", edgecolors="black", linewidth=0.5,
    )
    for _, row in eng.iterrows():
        ax.annotate(row["ticker"], (row["avg_score"], row["avg_comments"]),
                    fontsize=8, ha="center", va="bottom",
                    xytext=(0, 6), textcoords="offset points")
    ax.set_xlabel("Average Score (Upvotes)")
    ax.set_ylabel("Average Comments")
    ax.set_title("Engagement: Avg Score vs. Avg Comments per Coin\n"
                 "(bubble size = total post count)")
    plt.tight_layout()
    fig.savefig(os.path.join(plot_dir, "07_engagement_scatter.png"), dpi=150)
    plt.close(fig)

    print(f"[INFO] Saved 7 plots -> {plot_dir}/")

# ══════════════════════════════════════════════════════════════════════════════
# 9. MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Sentiment_Data_Load — Clean, preprocess, combine, and describe Reddit crypto data"
    )
    parser.add_argument("--raw_dir", type=str,
        default=os.path.join("Data", "Raw", "Sentiment"))
    parser.add_argument("--mapping", type=str,
        default="subreddit_ticker_mapping.xlsx")
    parser.add_argument("--output_csv", type=str,
        default=os.path.join("Data", "Processed", "Sentiment", "sentiment_combined.csv"))
    parser.add_argument("--output_stats", type=str,
        default=os.path.join("Outputs", "deskriptiv"))
    args = parser.parse_args(argv)

    # ── Path-shape tolerance ────────────────────────────────────────────
    # The script treats ``--output_stats`` as a *directory* (it appends the
    # XLSX file name internally at line ~564). Older CLI wrappers passed a
    # file path (``…/descriptive_statistics.xlsx``); in that case
    # ``os.makedirs(args.output_stats)`` would crash with FileExistsError
    # when the file already existed. Normalise here so both forms work.
    if args.output_stats.lower().endswith((".xlsx", ".xls", ".csv")):
        args.output_stats = os.path.dirname(args.output_stats) or "."

    if not os.path.isdir(args.raw_dir):
        print(f"[ERROR] Raw data directory not found: {args.raw_dir}")
        sys.exit(1)
    if not os.path.isfile(args.mapping):
        print(f"[ERROR] Mapping file not found: {args.mapping}")
        sys.exit(1)

    os.makedirs(os.path.dirname(args.output_csv), exist_ok=True)
    os.makedirs(args.output_stats, exist_ok=True)

    print("\n" + "=" * 60)
    print("STEP 1: Loading subreddit -> ticker mapping")
    print("=" * 60)
    mapping = load_ticker_mapping(args.mapping)

    print("\n" + "=" * 60)
    print("STEP 2: Loading raw sentiment data")
    print("=" * 60)
    df = load_raw_data(args.raw_dir, mapping)

    print("\n" + "=" * 60)
    print("STEP 3: Cleaning and deduplicating")
    print("=" * 60)
    df = clean_data(df)

    print("\n" + "=" * 60)
    print("STEP 4: Text preprocessing")
    print("=" * 60)
    df = apply_preprocessing(df)

    print("\n" + "=" * 60)
    print("STEP 5: Exporting processed data")
    print("=" * 60)

    export_cols = [
        "ticker", "subreddit", "date", "date_only", "hour",
        "removed", "upvote_ratio", "score", "total_awards_received",
        "num_comments", "num_crossposts", "title", "selftext",
    ]
    df[export_cols].to_csv(args.output_csv, index=False)
    file_mb = os.path.getsize(args.output_csv) / 1024 / 1024
    print(f"[INFO] Processed data -> {args.output_csv} ({file_mb:.1f} MB)")

    print("\n" + "=" * 60)
    print("STEP 6: Computing descriptive statistics")
    print("=" * 60)
    stats = compute_descriptive_stats(df)

    stats_path = os.path.join(args.output_stats, "descriptive_statistics.xlsx")
    export_stats_to_excel(stats, stats_path)

    print("\n" + "=" * 60)
    print("STEP 7: Creating visualizations")
    print("=" * 60)
    plot_dir = os.path.join(args.output_stats, "plots")
    create_visualizations(df, stats, plot_dir)

    print("\n" + "=" * 60)
    print("DONE")
    print("=" * 60)
    print(f"  Processed CSV:      {args.output_csv}")
    print(f"  Statistics Excel:   {stats_path}")
    print(f"  Plots:              {plot_dir}/")
    print()


if __name__ == "__main__":
    main()
