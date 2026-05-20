#!/usr/bin/env python3
# This wrapper is kept for backward compatibility. Preferred usage: python -m thesis_pipeline.cli stationarity
"""
Sentiment_Stationarity_Test_v2.py
==================================
Upgraded persistence and distributional-stability diagnostics for aggregated
Reddit sentiment features used as inputs to classification models.

Replaces the v1 pipeline (ADF + KPSS + Zivot-Andrews) with a journal-grade
test battery:

  Primary   : DF-GLS (Elliott-Rothenberg-Stock 1996) with BIC lag selection
              (MAIC is the academic ideal per Ng-Perron 2001, but `arch`
               does not expose MAIC; BIC is the parsimonious fallback).
  Robustness: Phillips-Perron (1988)
  Confirm.  : ADF and KPSS in both 'c' and 'ct' specifications
  Persistence: AR(1) rho-hat and half-life of shocks
  Long memory: Geweke-Porter-Hudak (1983) fractional-integration estimator
              (only computed when ADF and KPSS conflict)
  Breaks    : Bai-Perron (2003) multiple-break detection on the level
  Panel     : Pesaran (2007) CIPS test, one statistic per feature
  ML-shift  : Kolmogorov-Smirnov + Cramer-von Mises across an expanding
              time-series CV split

Multiple-testing: Benjamini-Hochberg FDR within (horizon x test) blocks.

Bug fixes vs v1:
  * Zivot-Andrews replaced by Bai-Perron; the v1 ZA call had two latent bugs
    ('method="AIC"' must be lowercase, 'za.breakpoint' is not a public attr)
    that silently produced zero ZA results.
  * 'classify_stationarity' rule corrected: 'both reject' under ('c','c')
    is conflicting evidence, NOT trend-stationary.
  * Bare except blocks replaced with logged, typed exception handling.
  * Minimum-n thresholds raised to 50 (ADF/KPSS/DF-GLS), 100 (Bai-Perron),
    200 (GPH).

Inputs:
    Data/Features/sentiment_features_6h.parquet
    Data/Features/sentiment_features_1h.parquet
    Data/Features/sentiment_features_1d.parquet

Outputs:
    Data/Features/stationarity_results_v2.xlsx
    Data/Features/plots_v2/*.png
    Data/Features/stationarity_records_v2.parquet  (raw, for replication)

Usage:
    python Sentiment_Stationarity_Test_v2.py
    python Sentiment_Stationarity_Test_v2.py --ticker BTC
    python Sentiment_Stationarity_Test_v2.py --horizon 1d --no_panel
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import time
from dataclasses import dataclass

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

import scipy.stats as sstats
from statsmodels.tsa.stattools import adfuller, kpss
from statsmodels.stats.multitest import multipletests
from statsmodels.regression.linear_model import OLS
from statsmodels.tools import add_constant
from statsmodels.tools.sm_exceptions import InterpolationWarning
from arch.unitroot import DFGLS, PhillipsPerron
import ruptures as rpt

import warnings
# KPSS hits its lookup-table boundary on most stationary series; the
# kpss_*_pval_clipped columns preserve which rows were clipped, so we
# silence the per-call warning to keep logs readable.
warnings.filterwarnings("ignore", category=InterpolationWarning)


# ──────────────────────────────────────────────────────────────────────
# 1. CONFIG
# ──────────────────────────────────────────────────────────────────────

FEATURE_DIR = os.path.join("Data", "Features")
PLOT_DIR    = os.path.join(FEATURE_DIR, "plots_v2")

HORIZONS = {
    "1h": "sentiment_features_1h.parquet",
    "6h": "sentiment_features_6h.parquet",
    "1d": "sentiment_features_1d.parquet",
}

SCORE_SUFFIXES = [
    "title_score_mean",
    "title_score_weighted_mean",
    "combined_score_mean",
    "combined_score_weighted_mean",
    "selftext_score_mean",
    "bullishness_ratio",
]
MODELS = ["vader", "finbert", "cryptobert"]
TOP_TICKERS = ["BTC", "ETH", "DOGE", "ADA", "SOL", "XRP"]

ALPHA = 0.05

MIN_N = {
    "adf":     50,
    "kpss":    50,
    "dfgls":   50,
    "pp":      50,
    "ar1":     30,
    "gph":     200,
    "bai":     100,
    "cips":    20,    # per-cross-section minimum; panel needs N>=4 sections too
    "ks_fold": 30,
}

# Known event dates relevant for crypto-sentiment in 2022
KNOWN_EVENTS = {
    "Terra/Luna": "2022-05-09",
    "FTX":        "2022-11-08",
}


# ──────────────────────────────────────────────────────────────────────
# 2. LOGGING
# ──────────────────────────────────────────────────────────────────────

logger = logging.getLogger("stationarity_v2")


def setup_logging(level: int = logging.INFO) -> None:
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    logging.basicConfig(level=level, format=fmt, datefmt="%H:%M:%S")
    # arch and statsmodels emit boundary warnings we want to count, not silence
    logging.getLogger("arch").setLevel(logging.ERROR)


# ──────────────────────────────────────────────────────────────────────
# 3. PRIMITIVES
# ──────────────────────────────────────────────────────────────────────

def _maxlag_schwert(n: int, k: int = 4) -> int:
    """
    Schwert (1989) data-dependent lag bound. k=4 is the parsimonious choice
    recommended by Ng & Perron (2001); k=12 is the original liberal bound.
    """
    return int(np.floor(k * (n / 100.0) ** 0.25))


def _safe_dropna(s: pd.Series) -> pd.Series:
    """
    Drop NaN values, but log an explicit warning if doing so introduces
    irregular spacing in what was a regular index. The series is returned
    as-is (gap-collapsed) for tests that cannot handle NaN; downstream
    interpretation must account for this.
    """
    if s.index.inferred_freq is None and s.isna().any():
        logger.debug("dropna on irregular index: %d NaNs removed", s.isna().sum())
    return s.dropna()


# ──────────────────────────────────────────────────────────────────────
# 4. UNIT-ROOT AND STATIONARITY TESTS
# ──────────────────────────────────────────────────────────────────────

def run_dfgls(series: pd.Series, trend: str = "c") -> dict:
    """
    Elliott-Rothenberg-Stock (1996) DF-GLS test. H0: unit root.
    Higher power than ADF in the near-unit-root region (Stock 1994).

    Lag selection: BIC. Ng & Perron (2001) recommend MAIC, but the `arch`
    package does not expose MAIC for DFGLS (only 'aic', 'bic', 't-stat').
    BIC is the standard fallback and is more parsimonious than AIC, which
    matters because over-selected lags inflate ADF/DF-GLS size in finite
    samples (Schwert 1989). For full MAIC, see Perron & Qu (2007) or
    implement Ng-Perron (2001) eq. (16) manually.
    """
    s = _safe_dropna(series)
    out = {f"dfgls_{trend}_stat": np.nan, f"dfgls_{trend}_pval": np.nan,
           f"dfgls_{trend}_lags": np.nan, f"dfgls_{trend}_reject": np.nan}
    if len(s) < MIN_N["dfgls"]:
        return out
    try:
        test = DFGLS(s.values, trend=trend, method="bic",
                     max_lags=_maxlag_schwert(len(s)))
        out[f"dfgls_{trend}_stat"]   = float(test.stat)
        out[f"dfgls_{trend}_pval"]   = float(test.pvalue)
        out[f"dfgls_{trend}_lags"]   = int(test.lags)
        out[f"dfgls_{trend}_reject"] = bool(test.pvalue < ALPHA)
    except (ValueError, np.linalg.LinAlgError) as e:
        logger.debug("DF-GLS failed on n=%d trend=%s: %s", len(s), trend, e)
    return out


def run_pp(series: pd.Series, trend: str = "c") -> dict:
    """
    Phillips-Perron (1988) test. Non-parametric correction for serial
    correlation; used as robustness check against ADF lag specification.
    H0: unit root.
    """
    s = _safe_dropna(series)
    out = {f"pp_{trend}_stat": np.nan, f"pp_{trend}_pval": np.nan,
           f"pp_{trend}_reject": np.nan}
    if len(s) < MIN_N["pp"]:
        return out
    try:
        test = PhillipsPerron(s.values, trend=trend)
        out[f"pp_{trend}_stat"]   = float(test.stat)
        out[f"pp_{trend}_pval"]   = float(test.pvalue)
        out[f"pp_{trend}_reject"] = bool(test.pvalue < ALPHA)
    except (ValueError, np.linalg.LinAlgError) as e:
        logger.debug("PP failed on n=%d trend=%s: %s", len(s), trend, e)
    return out


def run_adf(series: pd.Series, trend: str = "c") -> dict:
    """
    Augmented Dickey-Fuller. BIC lag selection (more parsimonious than AIC,
    less prone to size distortion in the presence of MA components).
    Maxlag bounded by parsimonious Schwert rule (k=4).
    H0: unit root.
    """
    s = _safe_dropna(series)
    out = {f"adf_{trend}_stat": np.nan, f"adf_{trend}_pval": np.nan,
           f"adf_{trend}_lags": np.nan, f"adf_{trend}_reject": np.nan}
    if len(s) < MIN_N["adf"]:
        return out
    try:
        stat, pval, lags, _, _, _ = adfuller(
            s.values, autolag="BIC",
            maxlag=_maxlag_schwert(len(s)),
            regression=trend,
        )
        out[f"adf_{trend}_stat"]   = float(stat)
        out[f"adf_{trend}_pval"]   = float(pval)
        out[f"adf_{trend}_lags"]   = int(lags)
        out[f"adf_{trend}_reject"] = bool(pval < ALPHA)
    except (ValueError, np.linalg.LinAlgError) as e:
        logger.debug("ADF failed on n=%d trend=%s: %s", len(s), trend, e)
    return out


def run_kpss(series: pd.Series, trend: str = "c") -> dict:
    """
    KPSS (Kwiatkowski-Phillips-Schmidt-Shin 1992). H0: stationary.
    Returns the test statistic alongside the p-value because statsmodels
    clips reported p-values to [0.01, 0.10]; the statistic itself permits
    a finer-grained decision.
    """
    s = _safe_dropna(series)
    out = {f"kpss_{trend}_stat": np.nan, f"kpss_{trend}_pval": np.nan,
           f"kpss_{trend}_lags": np.nan, f"kpss_{trend}_reject": np.nan,
           f"kpss_{trend}_pval_clipped": np.nan}
    if len(s) < MIN_N["kpss"]:
        return out
    try:
        stat, pval, lags, _ = kpss(s.values, regression=trend, nlags="auto")
        out[f"kpss_{trend}_stat"]   = float(stat)
        out[f"kpss_{trend}_pval"]   = float(pval)
        out[f"kpss_{trend}_lags"]   = int(lags)
        out[f"kpss_{trend}_reject"] = bool(pval < ALPHA)
        # boundary indicator: pval == 0.01 or 0.10 means we hit the table edge
        out[f"kpss_{trend}_pval_clipped"] = bool(pval in (0.01, 0.10))
    except (ValueError, np.linalg.LinAlgError) as e:
        logger.debug("KPSS failed on n=%d trend=%s: %s", len(s), trend, e)
    return out


# ──────────────────────────────────────────────────────────────────────
# 5. PERSISTENCE: AR(1) rho and half-life
# ──────────────────────────────────────────────────────────────────────

def run_ar1_persistence(series: pd.Series) -> dict:
    """
    Estimates the AR(1) coefficient by OLS on the demeaned series and
    derives the half-life of a shock as -ln(2)/ln(rho_hat). For rho>=1 or
    rho<=0 the half-life is reported as NaN; for highly persistent series
    the half-life is the most readable single persistence statistic.
    """
    s = _safe_dropna(series)
    out = {"ar1_rho": np.nan, "ar1_se": np.nan, "ar1_halflife": np.nan,
           "ar1_n": int(len(s))}
    if len(s) < MIN_N["ar1"]:
        return out
    try:
        y    = s.values
        y_t  = y[1:]
        y_l  = y[:-1]
        X    = add_constant(y_l)
        res  = OLS(y_t, X).fit(cov_type="HAC", cov_kwds={"maxlags": int(np.ceil(len(s) ** (1/3)))})
        rho  = float(res.params[1])
        se   = float(res.bse[1])
        out["ar1_rho"] = rho
        out["ar1_se"]  = se
        if 0.0 < rho < 1.0:
            out["ar1_halflife"] = -np.log(2) / np.log(rho)
    except (ValueError, np.linalg.LinAlgError) as e:
        logger.debug("AR(1) failed on n=%d: %s", len(s), e)
    return out


# ──────────────────────────────────────────────────────────────────────
# 6. LONG MEMORY: GPH log-periodogram
# ──────────────────────────────────────────────────────────────────────

def run_gph(series: pd.Series, m_power: float = 0.5) -> dict:
    """
    Geweke-Porter-Hudak (1983) log-periodogram regression for the
    fractional-integration parameter d. Uses m = floor(n^m_power)
    Fourier ordinates. d in (-0.5, 0.5) implies stationarity;
    d in [0.5, 1) implies non-stationary but mean-reverting; d>=1 unit-root.

    Reported standard error is the asymptotic OLS-corrected SE pi/sqrt(24*m).
    """
    s = _safe_dropna(series)
    out = {"gph_d": np.nan, "gph_se": np.nan, "gph_pval_d0": np.nan,
           "gph_m": np.nan, "gph_n": int(len(s))}
    if len(s) < MIN_N["gph"]:
        return out
    try:
        x = s.values - np.mean(s.values)
        n = len(x)
        m = int(np.floor(n ** m_power))
        if m < 5:
            return out
        # periodogram at Fourier frequencies lambda_j = 2*pi*j/n, j=1..m
        freqs = 2 * np.pi * np.arange(1, m + 1) / n
        # FFT periodogram
        ft = np.fft.fft(x)
        pgram = (np.abs(ft) ** 2) / (2 * np.pi * n)
        # match to first m positive frequencies
        I = pgram[1:m + 1]
        y = np.log(I)
        X = -np.log(4 * np.sin(freqs / 2) ** 2)
        # OLS: y = beta_0 + d * X
        Xc = add_constant(X)
        res = OLS(y, Xc).fit()
        d_hat = float(res.params[1])
        se_asy = np.pi / np.sqrt(24.0 * m)   # asymptotic SE
        # H0: d=0 vs H1: d!=0 (note: standard literature reports one-sided H1: d>0)
        z = d_hat / se_asy
        pval = 2.0 * (1 - sstats.norm.cdf(abs(z)))
        out["gph_d"]      = d_hat
        out["gph_se"]     = se_asy
        out["gph_pval_d0"] = float(pval)
        out["gph_m"]      = int(m)
    except (ValueError, np.linalg.LinAlgError) as e:
        logger.debug("GPH failed on n=%d: %s", len(s), e)
    return out


# ──────────────────────────────────────────────────────────────────────
# 7. STRUCTURAL BREAKS: Bai-Perron via binary segmentation
# ──────────────────────────────────────────────────────────────────────

def run_breaks(series: pd.Series, n_breaks: int = 2,
               min_size: int = 30) -> dict:
    """
    Detects up to n_breaks change points in mean using the Bai-Perron
    framework, implemented here via the `ruptures` package (binary
    segmentation with l2 cost). Returns the change-point dates and the
    segment means.

    For the daily horizon with the 2022 sample we expect at most two
    breaks (Terra/Luna May 2022, FTX November 2022), hence n_breaks=2.
    """
    s = _safe_dropna(series)
    out = {"n_breaks": np.nan, "break_dates": [], "segment_means": []}
    if len(s) < MIN_N["bai"] or len(s) < n_breaks * min_size + min_size:
        return out
    try:
        algo = rpt.Binseg(model="l2", min_size=min_size, jump=1).fit(s.values)
        # ruptures returns a list of break indices INCLUDING the endpoint
        bkps = algo.predict(n_bkps=n_breaks)
        # Convert indices to timestamps; drop the trailing endpoint
        idx_breaks = bkps[:-1]
        dates = [str(s.index[i].date()) if hasattr(s.index[i], "date") else str(s.index[i])
                 for i in idx_breaks]
        # segment means
        segments = np.split(s.values, idx_breaks)
        means = [float(np.mean(seg)) for seg in segments]
        out["n_breaks"]       = int(len(idx_breaks))
        out["break_dates"]    = dates
        out["segment_means"]  = means
    except (ValueError, IndexError) as e:
        logger.debug("Breaks failed on n=%d: %s", len(s), e)
    return out


# ──────────────────────────────────────────────────────────────────────
# 8. PANEL: CIPS (Pesaran 2007)
# ──────────────────────────────────────────────────────────────────────

def run_cips(panel: pd.DataFrame, trend: str = "c",
             max_lags: int = 4) -> dict:
    """
    Pesaran (2007) cross-sectionally augmented IPS test.

    panel: DataFrame indexed by timestamp, columns are tickers, values are
           the feature of interest (one feature at a time).
    trend: 'c' (constant) or 'ct' (constant + trend) regression for CADF.

    Algorithm:
      For each cross-section i, run the CADF regression
        Delta y_{i,t} = a_i + b_i * y_{i,t-1} + c_i * y_bar_{t-1}
                        + d_i * Delta y_bar_t
                        + sum_{k=1..p} g_{i,k} * Delta y_{i,t-k}
                        + (theta_i * t)?
                        + eps_{i,t}
      where y_bar_t is the cross-sectional mean.
      The CIPS statistic is the average of the t-statistics on b_i.

    Critical values are from Pesaran (2007) Table 2 (large N, T).
    """
    out = {"cips_stat": np.nan, "n_sections": np.nan, "n_obs": np.nan,
           "cips_5pct_cv": np.nan, "cips_reject": np.nan,
           "individual_tstats": []}

    # Cross-sectional mean y_bar at each timestamp (skip NaN per row)
    panel = panel.copy().dropna(how="all")
    if panel.shape[1] < 4 or panel.shape[0] < MIN_N["cips"] + max_lags + 2:
        return out

    y_bar    = panel.mean(axis=1, skipna=True)
    dy_bar   = y_bar.diff()
    y_bar_l  = y_bar.shift(1)

    tstats = []
    for col in panel.columns:
        y = panel[col]
        # require enough non-missing obs in this section
        valid = y.dropna()
        if len(valid) < MIN_N["cips"] + max_lags + 2:
            continue
        dy   = y.diff()
        y_l  = y.shift(1)
        # build design matrix
        df = pd.DataFrame({
            "dy":      dy,
            "y_l":     y_l,
            "y_bar_l": y_bar_l,
            "dy_bar":  dy_bar,
        })
        for k in range(1, max_lags + 1):
            df[f"dy_l{k}"] = dy.shift(k)
        if trend == "ct":
            df["t"] = np.arange(len(df))
        df = df.dropna()
        if len(df) < 20:
            continue
        try:
            X_cols = [c for c in df.columns if c != "dy"]
            X = add_constant(df[X_cols].values)
            res = OLS(df["dy"].values, X).fit()
            # b_i is the coefficient on y_l, which is column index 1+(X_cols.index('y_l'))
            j = X_cols.index("y_l") + 1   # +1 for the constant
            tstats.append(float(res.tvalues[j]))
        except (ValueError, np.linalg.LinAlgError):
            continue

    if len(tstats) < 4:
        return out

    # Truncate at -6.19 / +2.61 per Pesaran 2007 to bound moments
    trunc = np.clip(tstats, -6.19, 2.61)
    cips = float(np.mean(trunc))

    # Approximate 5% critical values from Pesaran 2007 Table 2(b)
    # (large N, T -> -2.15 for trend=c, -2.66 for trend=ct).
    cv5 = -2.15 if trend == "c" else -2.66

    out["cips_stat"]         = cips
    out["n_sections"]        = int(len(tstats))
    out["n_obs"]             = int(panel.shape[0])
    out["cips_5pct_cv"]      = cv5
    out["cips_reject"]       = bool(cips < cv5)
    out["individual_tstats"] = tstats
    return out


# ──────────────────────────────────────────────────────────────────────
# 9. JOINT CLASSIFICATION (corrected from v1)
# ──────────────────────────────────────────────────────────────────────

def classify_joint(adf_c_reject, adf_ct_reject,
                   kpss_c_reject, kpss_ct_reject) -> str:
    """
    Joint ADF + KPSS interpretation under both 'c' and 'ct' specifications,
    following Maddala & Kim (1998).

    Returns one of:
      - 'I(0)'                : level-stationary
      - 'trend-stationary'    : stationary around deterministic trend
      - 'I(1)'                : unit-root (non-stationary)
      - 'conflicting'         : ADF and KPSS contradict each other,
                                indicates structural break, fractional
                                integration, or test misspecification
      - 'inconclusive'        : neither test rejects (low power)
      - 'n/a'                 : insufficient data
    """
    flags = (adf_c_reject, adf_ct_reject, kpss_c_reject, kpss_ct_reject)
    if any(pd.isna(f) for f in flags):
        return "n/a"

    # Trend-stationary diagnostic uses 'ct' specification
    if adf_ct_reject and not kpss_ct_reject:
        # but if 'c' tests already agree, prefer the simpler classification
        if adf_c_reject and not kpss_c_reject:
            return "I(0)"
        return "trend-stationary"

    # Level-stationary diagnostic uses 'c' specification
    if adf_c_reject and not kpss_c_reject:
        return "I(0)"

    # Non-stationary
    if not adf_c_reject and kpss_c_reject:
        return "I(1)"

    # Conflicting
    if adf_c_reject and kpss_c_reject:
        # but we already caught the trend-stationary case above
        return "conflicting"

    # Neither rejects
    return "inconclusive"


# ──────────────────────────────────────────────────────────────────────
# 10. MAIN PER-SERIES BATTERY
# ──────────────────────────────────────────────────────────────────────

def test_series(series: pd.Series, run_long_memory_default: bool = False) -> dict:
    """
    Runs the full test battery on a single series. GPH is computed only
    when ADF and KPSS produce conflicting evidence, unless
    run_long_memory_default=True (used for top tickers at daily horizon).
    """
    rec: dict = {"n_obs": int(len(series))}
    # ── Primary
    rec.update(run_dfgls(series, trend="c"))
    rec.update(run_dfgls(series, trend="ct"))
    # ── Robustness
    rec.update(run_pp(series, trend="c"))
    rec.update(run_pp(series, trend="ct"))
    # ── Confirmation (both specifications)
    rec.update(run_adf(series, trend="c"))
    rec.update(run_adf(series, trend="ct"))
    rec.update(run_kpss(series, trend="c"))
    rec.update(run_kpss(series, trend="ct"))
    # ── Persistence
    rec.update(run_ar1_persistence(series))
    # ── Joint classification (uses confirmation tests for the verdict)
    rec["joint_class"] = classify_joint(
        rec.get("adf_c_reject"),  rec.get("adf_ct_reject"),
        rec.get("kpss_c_reject"), rec.get("kpss_ct_reject"),
    )
    # ── Long memory (conditional or forced)
    do_gph = run_long_memory_default or rec["joint_class"] == "conflicting"
    if do_gph:
        rec.update(run_gph(series))
    else:
        rec.update({k: np.nan for k in
                    ("gph_d", "gph_se", "gph_pval_d0", "gph_m", "gph_n")})
    return rec


def test_horizon(df: pd.DataFrame, horizon: str,
                 tickers: list | None = None) -> list[dict]:
    """
    Iterates over (ticker x feature) and runs the full battery.
    Bai-Perron is run only on daily data, top tickers, on title_score_mean
    columns (the natural break-detection target).
    """
    feature_cols = _get_feature_columns(df)
    if tickers is None:
        tickers = sorted(df["ticker"].unique())

    records = []
    total = len(tickers) * len(feature_cols)
    done = 0
    t0 = time.time()

    for ticker in tickers:
        sub = (df[df["ticker"] == ticker]
               .sort_values("timestamp")
               .set_index("timestamp"))
        # log post_count for use in volume-quintile stratification later
        post_count_total = int(sub["post_count"].sum()) if "post_count" in sub.columns else np.nan

        for feat in feature_cols:
            done += 1
            if done % 50 == 0:
                elapsed = time.time() - t0
                logger.info("  %s: %d/%d (%.1fs elapsed)", horizon, done, total, elapsed)

            series = sub[feat]

            base = {
                "horizon": horizon,
                "ticker":  ticker,
                "feature": feat,
                "post_count_total": post_count_total,
                "mean":    float(series.dropna().mean()) if series.dropna().size else np.nan,
                "std":     float(series.dropna().std())  if series.dropna().size else np.nan,
            }

            n_clean = series.dropna().size
            if n_clean < min(MIN_N["adf"], MIN_N["dfgls"]):
                base.update({"joint_class": "n/a", "n_obs": int(n_clean)})
                records.append(base)
                continue

            try:
                rec = test_series(series, run_long_memory_default=(horizon == "1d" and ticker in TOP_TICKERS))
            except Exception as e:    # last-resort guard, but log
                logger.exception("Battery crashed on %s/%s/%s: %s", horizon, ticker, feat, e)
                rec = {"joint_class": "error", "n_obs": int(n_clean)}

            base.update(rec)

            # Bai-Perron: daily, top tickers only, on the score (not bullishness)
            if (horizon == "1d" and ticker in TOP_TICKERS
                and feat.endswith("_title_score_mean")):
                br = run_breaks(series, n_breaks=2)
                base["n_breaks"]      = br["n_breaks"]
                base["break_dates"]   = ";".join(br["break_dates"]) if br["break_dates"] else ""
                base["segment_means"] = ";".join(f"{m:.4f}" for m in br["segment_means"]) if br["segment_means"] else ""

            records.append(base)

    return records


def _get_feature_columns(df: pd.DataFrame) -> list[str]:
    cols = []
    for model in MODELS:
        for suffix in SCORE_SUFFIXES:
            col = f"{model}_{suffix}"
            if col in df.columns:
                cols.append(col)
    if "post_count" in df.columns:
        cols.append("post_count")
    return cols


# ──────────────────────────────────────────────────────────────────────
# 11. MULTIPLE-TESTING CORRECTION
# ──────────────────────────────────────────────────────────────────────

def apply_fdr(records: pd.DataFrame, pval_cols: list[str], alpha: float = ALPHA) -> pd.DataFrame:
    """
    Benjamini-Hochberg FDR correction within each (horizon x test) block.
    Adds *_pval_bh and *_reject_bh columns.
    """
    out = records.copy()
    for hz, sub in records.groupby("horizon"):
        for pc in pval_cols:
            if pc not in sub.columns:
                continue
            mask = sub[pc].notna()
            if mask.sum() < 2:
                continue
            pvals = sub.loc[mask, pc].values
            reject, pvals_corr, _, _ = multipletests(pvals, alpha=alpha, method="fdr_bh")
            out.loc[sub.index[mask], f"{pc}_bh"]    = pvals_corr
            out.loc[sub.index[mask], f"{pc.replace('pval','reject')}_bh"] = reject
    return out


# ──────────────────────────────────────────────────────────────────────
# 12. PANEL TESTS — one CIPS per feature per horizon
# ──────────────────────────────────────────────────────────────────────

def run_all_panels(df: pd.DataFrame, horizon: str) -> list[dict]:
    """
    For each feature column, build a (timestamp x ticker) panel and run CIPS
    in both 'c' and 'ct' specifications.
    """
    feature_cols = _get_feature_columns(df)
    out = []
    for feat in feature_cols:
        wide = (df.pivot_table(index="timestamp", columns="ticker", values=feat,
                               aggfunc="first")
                  .sort_index())
        for trend in ("c", "ct"):
            cips = run_cips(wide, trend=trend)
            out.append({
                "horizon": horizon,
                "feature": feat,
                "trend":   trend,
                **{k: v for k, v in cips.items() if k != "individual_tstats"},
            })
    return out


# ──────────────────────────────────────────────────────────────────────
# 13. TIME-SERIES CV STABILITY (KS, Cramer-von Mises)
# ──────────────────────────────────────────────────────────────────────

def fold_stability(series: pd.Series, n_splits: int = 5) -> list[dict]:
    """
    Expanding-window TS-CV: split the index into n_splits+1 folds, then
    for each fold k compute KS and CvM between fold k (train) and fold k+1
    (validation). Returns one record per split, per series.
    """
    s = _safe_dropna(series)
    out = []
    if len(s) < MIN_N["ks_fold"] * (n_splits + 1):
        return out
    fold_size = len(s) // (n_splits + 1)
    for k in range(n_splits):
        train_end = (k + 1) * fold_size
        val_end   = (k + 2) * fold_size
        train = s.iloc[:train_end].values
        val   = s.iloc[train_end:val_end].values
        if len(train) < MIN_N["ks_fold"] or len(val) < MIN_N["ks_fold"]:
            continue
        try:
            ks_stat, ks_p = sstats.ks_2samp(train, val, alternative="two-sided")
            cvm = sstats.cramervonmises_2samp(train, val)
            out.append({
                "split":     k,
                "n_train":   len(train),
                "n_val":     len(val),
                "ks_stat":   float(ks_stat),
                "ks_pval":   float(ks_p),
                "cvm_stat":  float(cvm.statistic),
                "cvm_pval":  float(cvm.pvalue),
            })
        except (ValueError, RuntimeError) as e:
            logger.debug("Fold stability failed: %s", e)
    return out


def run_all_fold_stability(df: pd.DataFrame, horizon: str,
                           tickers: list[str] | None = None) -> list[dict]:
    """Daily horizon only (the only one with enough fold size)."""
    if horizon != "1d":
        return []
    feature_cols = _get_feature_columns(df)
    if tickers is None:
        tickers = TOP_TICKERS
    out = []
    for ticker in tickers:
        sub = (df[df["ticker"] == ticker]
               .sort_values("timestamp")
               .set_index("timestamp"))
        for feat in feature_cols:
            for rec in fold_stability(sub[feat]):
                rec.update({"horizon": horizon, "ticker": ticker, "feature": feat})
                out.append(rec)
    return out


# ──────────────────────────────────────────────────────────────────────
# 14. VOLUME STRATIFICATION
# ──────────────────────────────────────────────────────────────────────

def stratify_by_volume(records_df: pd.DataFrame) -> pd.DataFrame:
    """
    Within each horizon, assign each ticker to a volume quintile based
    on total post_count, and add a 'volume_quintile' column.
    """
    out = records_df.copy()
    out["volume_quintile"] = np.nan
    for hz, sub in records_df.groupby("horizon"):
        ticker_volume = sub.groupby("ticker")["post_count_total"].first().sort_values()
        try:
            qs = pd.qcut(ticker_volume, q=5, labels=[1, 2, 3, 4, 5], duplicates="drop")
            for ticker, q in qs.items():
                out.loc[(out["horizon"] == hz) & (out["ticker"] == ticker), "volume_quintile"] = int(q)
        except ValueError:
            # Not enough unique values; skip stratification at this horizon
            logger.warning("  %s: not enough unique post_count values for quintile split", hz)
    return out


# ──────────────────────────────────────────────────────────────────────
# 15. PLOTS — distribution-of-statistics, AR(1), half-life
# ──────────────────────────────────────────────────────────────────────

def plot_test_stat_distributions(records_df: pd.DataFrame, plot_dir: str) -> None:
    """
    For each horizon, plot histograms of:
      - DF-GLS test statistic (with 5% CV)
      - ADF (c) test statistic (with 5% CV)
      - KPSS (c) test statistic (with 5% CV)
      - AR(1) rho-hat
      - Half-life (capped at 365 days for visualisation)
    """
    os.makedirs(plot_dir, exist_ok=True)

    for hz, sub in records_df.groupby("horizon"):
        fig, axes = plt.subplots(2, 3, figsize=(15, 8))
        axes = axes.ravel()

        # DF-GLS
        ax = axes[0]
        ax.hist(sub["dfgls_c_stat"].dropna(), bins=40, color="#2196F3", edgecolor="white")
        ax.axvline(-1.95, color="red", linestyle="--", label="5% CV (~-1.95)")
        ax.set_title(f"{hz} — DF-GLS (c) statistic")
        ax.set_xlabel("DF-GLS stat"); ax.legend(); ax.grid(alpha=0.2)

        # ADF c
        ax = axes[1]
        ax.hist(sub["adf_c_stat"].dropna(), bins=40, color="#4CAF50", edgecolor="white")
        ax.axvline(-2.86, color="red", linestyle="--", label="5% CV (~-2.86)")
        ax.set_title(f"{hz} — ADF (c) statistic")
        ax.set_xlabel("ADF stat"); ax.legend(); ax.grid(alpha=0.2)

        # KPSS c
        ax = axes[2]
        ax.hist(sub["kpss_c_stat"].dropna(), bins=40, color="#FF9800", edgecolor="white")
        ax.axvline(0.463, color="red", linestyle="--", label="5% CV (0.463)")
        ax.set_title(f"{hz} — KPSS (c) statistic")
        ax.set_xlabel("KPSS stat"); ax.legend(); ax.grid(alpha=0.2)

        # AR(1) rho
        ax = axes[3]
        ax.hist(sub["ar1_rho"].dropna(), bins=40, color="#9C27B0", edgecolor="white")
        ax.axvline(1.0, color="red", linestyle="--", label="ρ=1 (unit root)")
        ax.axvline(0.0, color="grey", linestyle=":", alpha=0.5, label="ρ=0 (white noise)")
        ax.set_title(f"{hz} — AR(1) ρ̂")
        ax.set_xlabel("ρ̂"); ax.legend(); ax.grid(alpha=0.2)

        # Half-life (capped)
        ax = axes[4]
        hl = sub["ar1_halflife"].dropna()
        hl_cap = hl[hl > 0].clip(upper=365)
        ax.hist(hl_cap, bins=40, color="#00BCD4", edgecolor="white")
        ax.set_title(f"{hz} — Halbwertszeit Schock (Tage / Bins, capped 365)")
        ax.set_xlabel("Half-life"); ax.grid(alpha=0.2)

        # Lag selection (DF-GLS)
        ax = axes[5]
        ax.hist(sub["dfgls_c_lags"].dropna(), bins=range(0, 30), color="#607D8B", edgecolor="white")
        ax.set_title(f"{hz} — DF-GLS gewählte Lags (MAIC)")
        ax.set_xlabel("Lags"); ax.grid(alpha=0.2)

        fig.tight_layout()
        fname = os.path.join(plot_dir, f"distribution_stats_{hz}.png")
        fig.savefig(fname, dpi=140, bbox_inches="tight")
        plt.close(fig)
        logger.info("  saved %s", fname)


def plot_rolling_stats(df_daily: pd.DataFrame, plot_dir: str,
                       tickers: list | None = None, window: int = 90) -> None:
    """
    Rolling mean and rolling standard deviation of title_score_mean across
    the three sentiment models. Known crypto events are marked as vertical
    lines.

    Fix:
      - Ensures timestamp column is timezone-aware UTC.
      - Event dates are created as UTC timestamps.
      - Prevents tz-naive vs tz-aware comparison errors.
    """
    if tickers is None:
        tickers = TOP_TICKERS

    os.makedirs(plot_dir, exist_ok=True)

    df_daily = df_daily.copy()
    df_daily["timestamp"] = pd.to_datetime(df_daily["timestamp"], utc=True)

    colors = {
        "vader": "#2196F3",
        "finbert": "#4CAF50",
        "cryptobert": "#FF9800",
    }

    for ticker in tickers:
        sub = df_daily[df_daily["ticker"] == ticker].sort_values("timestamp")

        if len(sub) < window + 10:
            continue

        fig, axes = plt.subplots(2, 1, figsize=(14, 7), sharex=True)

        for model in MODELS:
            col = f"{model}_title_score_mean"
            if col not in sub.columns:
                continue

            s = sub.set_index("timestamp")[col].dropna()

            axes[0].plot(
                s.rolling(window).mean(),
                label=model.upper(),
                color=colors[model],
                linewidth=1.2,
            )

            axes[1].plot(
                s.rolling(window).std(),
                label=model.upper(),
                color=colors[model],
                linewidth=1.2,
            )

        axes[0].set_title(f"{ticker} — {window}-Tage Rolling Mean (Title Score)")
        axes[0].axhline(
            0,
            color="black",
            linewidth=0.5,
            linestyle="--",
            alpha=0.4,
        )
        axes[0].set_ylabel("Rolling Mean")
        axes[0].legend(loc="upper right")
        axes[0].grid(alpha=0.2)

        axes[1].set_title(f"{ticker} — {window}-Tage Rolling Std")
        axes[1].set_ylabel("Rolling Std")
        axes[1].legend(loc="upper right")
        axes[1].grid(alpha=0.2)

        axes[1].xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        axes[1].xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
        plt.xticks(rotation=45)

        ts_min = sub["timestamp"].min()
        ts_max = sub["timestamp"].max()

        for label, date_str in KNOWN_EVENTS.items():
            dt = pd.Timestamp(date_str, tz="UTC")

            if ts_min <= dt <= ts_max:
                for ax in axes:
                    ax.axvline(
                        dt,
                        color="red",
                        linewidth=0.8,
                        linestyle="--",
                        alpha=0.6,
                    )

                axes[0].text(
                    dt,
                    axes[0].get_ylim()[1] * 0.95,
                    f" {label}",
                    fontsize=8,
                    color="red",
                    verticalalignment="top",
                )

        plt.tight_layout()

        fname = os.path.join(plot_dir, f"rolling_stats_{ticker}.png")
        plt.savefig(fname, dpi=140, bbox_inches="tight")
        plt.close(fig)

        logger.info("  saved %s", fname)


# ──────────────────────────────────────────────────────────────────────
# 16. EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────

def save_excel(records_df: pd.DataFrame, panel_df: pd.DataFrame,
               fold_df: pd.DataFrame, output_path: str) -> None:
    """Write multi-sheet workbook with per-horizon results, panel CIPS, and CV stability."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); wb.remove(wb.active)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    data_font   = Font(name="Arial", size=10)
    thin_bot    = Border(bottom=Side(style="thin", color="D9D9D9"))
    fills = {
        "I(0)":             PatternFill("solid", fgColor="C6EFCE"),
        "trend-stationary": PatternFill("solid", fgColor="FFEB9C"),
        "I(1)":             PatternFill("solid", fgColor="FFC7CE"),
        "conflicting":      PatternFill("solid", fgColor="FCE4D6"),
        "inconclusive":     PatternFill("solid", fgColor="EDEDED"),
    }

    # ── Summary sheet ───────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    summary_rows = []
    for hz in ["1h", "6h", "1d"]:
        sub = records_df[records_df["horizon"] == hz]
        for cls in ["I(0)", "trend-stationary", "I(1)", "conflicting", "inconclusive", "n/a"]:
            n = (sub["joint_class"] == cls).sum()
            pct = 100 * n / len(sub) if len(sub) else 0.0
            summary_rows.append({"Horizon": hz, "Class": cls, "Count": n, "Pct": pct})
    summary_rows.append({"Horizon": "", "Class": "", "Count": "", "Pct": ""})
    for hz in ["1h", "6h", "1d"]:
        sub = records_df[records_df["horizon"] == hz]
        for col in ["ar1_rho", "ar1_halflife"]:
            if col not in sub.columns: continue
            v = sub[col].dropna()
            if len(v):
                summary_rows.append({
                    "Horizon": hz, "Class": f"{col} median",
                    "Count": "", "Pct": float(np.median(v)),
                })
    headers = ["Horizon", "Class", "Count", "Pct"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")
    for ri, rec in enumerate(summary_rows, 2):
        for ci, k in enumerate(headers, 1):
            v = rec[k]
            if isinstance(v, float): v = round(v, 4)
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = data_font; c.border = thin_bot
            if k == "Class" and rec["Class"] in fills:
                c.fill = fills[rec["Class"]]
    for ci in range(1, 5):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    # ── Per-horizon sheets ──────────────────────────────────────────
    cols_full = [
        "horizon", "ticker", "feature", "n_obs", "post_count_total", "volume_quintile",
        "mean", "std",
        "dfgls_c_stat", "dfgls_c_pval", "dfgls_c_pval_bh", "dfgls_c_lags", "dfgls_c_reject",
        "dfgls_ct_stat", "dfgls_ct_pval", "dfgls_ct_pval_bh", "dfgls_ct_reject",
        "pp_c_stat", "pp_c_pval", "pp_c_pval_bh", "pp_c_reject",
        "pp_ct_stat", "pp_ct_pval", "pp_ct_reject",
        "adf_c_stat", "adf_c_pval", "adf_c_pval_bh", "adf_c_lags", "adf_c_reject",
        "adf_ct_stat", "adf_ct_pval", "adf_ct_reject",
        "kpss_c_stat", "kpss_c_pval", "kpss_c_pval_bh", "kpss_c_reject", "kpss_c_pval_clipped",
        "kpss_ct_stat", "kpss_ct_pval", "kpss_ct_reject", "kpss_ct_pval_clipped",
        "ar1_rho", "ar1_se", "ar1_halflife",
        "gph_d", "gph_se", "gph_pval_d0", "gph_m",
        "n_breaks", "break_dates", "segment_means",
        "joint_class",
    ]
    display = {c: c for c in cols_full}
    for hz in ["1h", "6h", "1d"]:
        sub = records_df[records_df["horizon"] == hz]
        if sub.empty: continue
        ws = wb.create_sheet(hz)
        cols = [c for c in cols_full if c in sub.columns]
        for ci, name in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=display.get(name, name))
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        for ri, (_, rec) in enumerate(sub.iterrows(), 2):
            for ci, name in enumerate(cols, 1):
                v = rec.get(name)
                if isinstance(v, (np.floating, float)) and pd.notna(v): v = round(float(v), 6)
                elif isinstance(v, (np.bool_, bool)):                   v = bool(v)
                elif isinstance(v, pd.Timestamp):                       v = v.strftime("%Y-%m-%d") if pd.notna(v) else ""
                elif pd.isna(v):                                        v = ""
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = data_font; c.border = thin_bot
                if name == "joint_class" and v in fills: c.fill = fills[v]
        for ci, name in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(name) + 2, 12)
        ws.freeze_panes = "A2"

    # ── Panel CIPS sheet ────────────────────────────────────────────
    if not panel_df.empty:
        ws = wb.create_sheet("Panel_CIPS")
        cols = list(panel_df.columns)
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        for ri, (_, rec) in enumerate(panel_df.iterrows(), 2):
            for ci, name in enumerate(cols, 1):
                v = rec.get(name)
                if isinstance(v, (np.floating, float)) and pd.notna(v): v = round(float(v), 6)
                elif isinstance(v, (np.bool_, bool)):                   v = bool(v)
                elif pd.isna(v):                                        v = ""
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = data_font; c.border = thin_bot
        for ci, name in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(name) + 2, 14)
        ws.freeze_panes = "A2"

    # ── CV Fold Stability sheet ─────────────────────────────────────
    if not fold_df.empty:
        ws = wb.create_sheet("CV_Fold_Stability")
        cols = list(fold_df.columns)
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        for ri, (_, rec) in enumerate(fold_df.iterrows(), 2):
            for ci, name in enumerate(cols, 1):
                v = rec.get(name)
                if isinstance(v, (np.floating, float)) and pd.notna(v): v = round(float(v), 6)
                elif pd.isna(v):                                        v = ""
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = data_font; c.border = thin_bot
        for ci, name in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(name) + 2, 12)
        ws.freeze_panes = "A2"

    abs_path = os.path.abspath(output_path)
    wb.save(abs_path)
    logger.info("Results saved -> %s", abs_path)


# ──────────────────────────────────────────────────────────────────────
# 17. MAIN
# ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Stationarity v2 (DF-GLS, PP, ADF, KPSS, GPH, CIPS, KS-CV)")
    parser.add_argument("--feature_dir", type=str, default=FEATURE_DIR)
    parser.add_argument("--ticker",   type=str, default=None, help="restrict to one ticker")
    parser.add_argument("--horizon",  type=str, default=None, choices=["1h", "6h", "1d"])
    parser.add_argument("--no_panel", action="store_true", help="skip CIPS panel tests")
    parser.add_argument("--no_plots", action="store_true")
    parser.add_argument("--debug",    action="store_true")
    args = parser.parse_args()

    setup_logging(level=logging.DEBUG if args.debug else logging.INFO)

    # version pinning to log
    import statsmodels, arch
    logger.info("statsmodels=%s  arch=%s  ruptures=%s", statsmodels.__version__, arch.__version__, rpt.__version__)

    # validate inputs
    for hz, fname in HORIZONS.items():
        path = os.path.join(args.feature_dir, fname)
        if not os.path.isfile(path):
            logger.error("missing input file: %s", path); sys.exit(1)

    horizons_to_test = {args.horizon: HORIZONS[args.horizon]} if args.horizon else HORIZONS
    all_records  = []
    panel_recs   = []
    fold_recs    = []

    for hz, fname in horizons_to_test.items():
        logger.info("=" * 60)
        logger.info("HORIZON: %s", hz)
        logger.info("=" * 60)
        df = pd.read_parquet(os.path.join(args.feature_dir, fname))
        logger.info("loaded %d rows, %d tickers", len(df), df["ticker"].nunique())
        tickers = [args.ticker] if args.ticker else None

        all_records.extend(test_horizon(df, hz, tickers=tickers))

        if not args.no_panel:
            logger.info("running CIPS panel for %s ...", hz)
            panel_recs.extend(run_all_panels(df, hz))

        fold_recs.extend(run_all_fold_stability(df, hz))

    # ── DataFrames
    rec_df = pd.DataFrame(all_records)
    panel_df = pd.DataFrame(panel_recs)
    fold_df = pd.DataFrame(fold_recs)

    # ── FDR correction within each horizon, per test
    pval_cols = ["dfgls_c_pval", "dfgls_ct_pval",
                 "pp_c_pval",
                 "adf_c_pval",
                 "kpss_c_pval"]
    rec_df = apply_fdr(rec_df, pval_cols=pval_cols, alpha=ALPHA)

    # ── Volume stratification
    rec_df = stratify_by_volume(rec_df)

    # ── Save raw parquet for replication (preserves dtypes; xlsx loses bools)
    parquet_path = os.path.join(args.feature_dir, "stationarity_records_v2.parquet")
    rec_df_safe = rec_df.copy()
    for c in rec_df_safe.columns:
        if rec_df_safe[c].dtype == object:
            rec_df_safe[c] = rec_df_safe[c].astype(str)
    rec_df_safe.to_parquet(parquet_path)
    logger.info("Raw records saved -> %s", parquet_path)

    # ── Save Excel
    xlsx_path = os.path.join(args.feature_dir, "stationarity_results_v2.xlsx")
    save_excel(rec_df, panel_df, fold_df, xlsx_path)

    # ── Plots
    if not args.no_plots:
        os.makedirs(PLOT_DIR, exist_ok=True)
        plot_test_stat_distributions(rec_df, PLOT_DIR)
        # rolling stats only on daily
        if "1d" in horizons_to_test:
            df_daily = pd.read_parquet(os.path.join(args.feature_dir, HORIZONS["1d"]))
            plot_rolling_stats(df_daily, PLOT_DIR,
                               tickers=[args.ticker] if args.ticker else TOP_TICKERS)

    logger.info("DONE")


if __name__ == "__main__":
    main()
